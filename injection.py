# -----------------------------------------------
# Created by Erick Delgado June 8th 2023

# Dedicated to Owen and Jessica the grid slaying
# ----> WIZARDS <----
#          __
#          | \      _
#       ==='=='==  (o>
#          \++/   / )
# __.-------------^^-.__
#    \----.    .----/
#          \_/\|
#          / _ \
#         / /\ \
#       _/_/  \_\_
#      /_/      \_\
# -----------------------------------------------

import openpyxl

def update_pricing():
    # Open the Price Changes file
    price_changes_wb = openpyxl.load_workbook('Excel-Files/Pricing-Change.xlsx')
    price_changes_sheet = price_changes_wb['INS Details']

    # Prompt user for inputs
    cruise = input("Select the first cruise that the Pricing Date will take effect on: ")
    end_date = input("Enter the End Date: ")
    pricing_date = input("Enter the Pricing Date: ")

    # Find the cruise in row 4 of Price Changes
    cruise_column = None
    for column in range(1, price_changes_sheet.max_column + 1):
        if price_changes_sheet.cell(row=4, column=column).value == cruise:
            cruise_column = column
            break

    if cruise_column is None:
        print("Cruise not found in Price Changes.")
        return

    # Open the Pricing Grids file
    pricing_grids_wb = openpyxl.load_workbook('Excel-Files/Grid-of-doom.xlsx')
    pricing_grids_sheet = pricing_grids_wb['Insignia']

    # Find the column in Pricing Grids that corresponds to the cruise
    pricing_grids_column = None
    for column in range(1, pricing_grids_sheet.max_column + 1):
        if pricing_grids_sheet.cell(row=3, column=column).value == cruise:
            pricing_grids_column = column
            break

    if pricing_grids_column is None:
        print("Cruise not found in Pricing Grids.")
        return

    # Update prices and dates for the remaining cruises
    for column in range(cruise_column, price_changes_sheet.max_column + 1):
        # Get NEW AC IN and NEW AC OUT values
        new_ac_in = price_changes_sheet.cell(row=14, column=column).value
        new_ac_out = price_changes_sheet.cell(row=15, column=column).value

        # Update NEW AC IN and NEW AC OUT in Pricing Grids (shifted one cell/row upwards)
        pricing_grids_sheet.cell(row=17, column=pricing_grids_column).value = new_ac_in
        pricing_grids_sheet.cell(row=18, column=pricing_grids_column).value = new_ac_out

        # Check the type in Price Changes row 85
        pricing_type = price_changes_sheet.cell(row=85, column=column).value

        # If type is "N", extend the end date in Pricing Grids based on the tier
        if pricing_type == "N":
            current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value
            row_number = 132 + (current_tier - 1) * 42
            pricing_grids_sheet.cell(row=row_number-2, column=pricing_grids_column).value = end_date

        # If type is "F" or "P", update prices in Pricing Grids for the respective tier
        elif pricing_type in ("F", "P"):
            current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value
            row_number = 300 + current_tier
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date
            pricing_grids_sheet.cell(row=row_number-2, column=pricing_grids_column).value = pricing_date
            pricing_grids_sheet.cell(row=row_number-3, column=pricing_grids_column).value = 0

        # Move to the next column in Price Changes and Pricing Grids
        cruise_column += 1
        pricing_grids_column += 1

    # Save the changes
    price_changes_wb.save('Excel-Files/Pricing-Change.xlsx')
    pricing_grids_wb.save('Excel-Files/Grid-of-doom.xlsx')

    print("Pricing update completed successfully.")

# Call the function to perform the update
update_pricing()
