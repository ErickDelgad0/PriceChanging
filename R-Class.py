# Created by Erick Delgado June 15th 2023
# Formatted for R-Class Price Changes

# Dedicated to Owen and Jessica the grid slaying GODS

#             .-------.
#             |(~\o/~)|
#             ||\/X\/||
#             || \ / ||
#             ||o X o||
#             || / \ ||
#             ||/\X/\||
# .----------._)~   ~(_.----------.
# |\/)~~(\/\   (~\ /~)   /\/)~~(\/|
# |(X () X) >o  >-X-<  o< (X () X)|
# |/\)__(/\/  _(_/|\_)_  \/\)__(/\|
# '----------' )     ( '----------'
#             ||\/~\/||
#             || \o/ ||
#             |(~\ /~)|
#             ||\/X\/||
#             || \ / ||
#             ||  X  ||
#             ||\(/\/||
#             ||=)O(=||
#             ||/\/)\||
#             ||  X  ||
#             || / \ ||
#             ||/\X/\||
#             |(_/o\_)|
#             '._____.'
# -----------------------------------------------


import openpyxl

def ArgSelector():
    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")
    print("....................WARNING....................")
    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")

    print(" ")
    print("This script is intended to work on R-Class only")
    print("Ensure excel files are in the same folder as R-Class.py")

    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")

    PriceGridFile = input("Enter the file name of excel PRICING GRID: ")
    print(" ")
    PriceChangeFile = input("Enter the file name of excel PRICING CHANGES: ")
    print(" ")
    Ship = input("Enter Ship Abbrev (ex. INS, NAU, etc.): ")
    print(" ")
    pricing_date = input("Enter the Pricing Date: ")
    print(" ")
    end_date = input("Enter the End Date: ")
    print(" ")
    start_cruise = input("Select the first cruise that the Pricing Date will take effect on: ")
    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")

    ArgInfo = [PriceGridFile, PriceChangeFile, Ship, pricing_date, end_date, start_cruise]
    return ArgInfo

def ArgConfirmation(ArgInfo):
    phrases = [
        "1: Price grid file: ",
        "2: Price change file: ",
        "3: Ship Selected: ",
        "4: Pricing Date Selected: ",
        "5: First Cruise Price Effects",
        "6: End of Pricing Date: "
    ]
    
    print("...Check if given inputs are correct...")
    print(f"{phrases[0]} {ArgInfo[0]}: ")
    print(f"{phrases[1]} {ArgInfo[1]}: ")
    print(f"{phrases[2]} {ArgInfo[2]}: ")
    print(f"{phrases[3]} {ArgInfo[3]}: ")
    print(f"{phrases[4]} {ArgInfo[4]}: ")
    print(f"{phrases[5]} {ArgInfo[5]}: ")

    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")
    print("....................WARNING....................")
    print(" -  -  -  -  - -  -  -  -  - -  -  -  -  -  -  ")

    PriceGrid = input("Are the Input Arguments Above Correct (Y/N): ")

    if (PriceGrid == "Y") or (PriceGrid == "y"):
        print("Proceeding...")
        return ArgInfo
    else:
        value = int(input("Select value to change (1, 2, 3, 4, 5, 6): "))
        value = value - 1
        new_value = input(f"New {phrases[value]}")
        ArgInfo[value] = new_value
        ArgConfirmation(ArgInfo)

def Class_Execution(Ship, Info):
    R_class = ["INS", "NAU", "REG", "SIR"]
    O_class = ["MNA", "RVA"]
    A_class = ["VIS"]
    if Ship in R_class:
        return R_Class_Change(Info)
    elif Ship in O_class:
        return O_Class_Change(Info)
    elif Ship in A_class:
        return A_Class_Change(Info)
    else:
        print("Ship not found in classes")
        return

def R_Class_Change(Info):
    column, price_changes_sheet, pricing_grids_sheet, pricing_grids_column, end_date, pricing_date = Info[:6]

    # store pricing values
    AI_values = []
    row_start = 182
    row_end = 197

    for row in range(row_start, row_end + 1):
        AI_values.append(price_changes_sheet.cell(row=row, column=column).value)

    AI_OS, AI_VS, AI_PH1, AI_PH2, AI_PH3, AI_A1, AI_A2, AI_A3, AI_B1, AI_B2, AI_C1, AI_C2, AI_D, AI_S, AI_F, AI_G = AI_values[:16]

    # Check the type in Price Changes row 85
    pricing_type = price_changes_sheet.cell(row=85, column=column).value

    # Check the Air Credit Action
    AirCreditAction = price_changes_sheet.cell(row=16, column=column).value

    # Current Tier
    current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value

    # If type is "N", extend the end date in Pricing Grids based on the tier
    if pricing_type == "N" and AirCreditAction == "FLAT":
        try:
            row_number = 132 + (current_tier - 1) * 42

            # Extending end date only
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date
            
        except:
            print(f"Current Column {column} was unable to process...")

    # If type is "F" or "P", update prices in Pricing Grids for the respective tier
    elif pricing_type in ("F", "P"):
        try:
            row_number = (132 + (current_tier-1) * 42) + 42

            # Discounts, dates, and pricing applied
            values = [
                0, pricing_date, end_date,
                AI_OS, AI_VS, AI_PH1,
                AI_PH2, AI_PH3, AI_A1,
                AI_A2, AI_A3, AI_B1,
                AI_B2, AI_C1, AI_C2,
                AI_D, AI_S, AI_F, AI_G
            ]

            for i, value in enumerate(values):
                pricing_grids_sheet.cell(row=row_number + i, column=pricing_grids_column).value = value

            # Increase the tier of price by 1
            pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value = current_tier + 1

        except:
            print(f"Current Column {column} was unable to process...")

def O_Class_Change(Info):
    column, price_changes_sheet, pricing_grids_sheet, pricing_grids_column, end_date, pricing_date = Info[:6]

    # store pricing values
    AI_values = []
    row_start = 190
    row_end = 206

    for row in range(row_start, row_end + 1):
        AI_values.append(price_changes_sheet.cell(row=row, column=column).value)

    AI_OS, AI_VS, AI_OC, AI_PH1, AI_PH2, AI_PH3, AI_A1, AI_A2, AI_A3, AI_A4, AI_B1, AI_B2, AI_B3, AI_B4, AI_C, AI_F, AI_G = AI_values[:17]

    # Check the type in Price Changes row 85
    pricing_type = price_changes_sheet.cell(row=88, column=column).value

    # Check the Air Credit Action
    AirCreditAction = price_changes_sheet.cell(row=16, column=column).value

    # Current Tier
    current_tier = pricing_grids_sheet.cell(row=66, column=pricing_grids_column).value

    # If type is "N", extend the end date in Pricing Grids based on the tier
    if pricing_type == "N" and AirCreditAction == "FLAT":
        try:
            row_number = 136 + (current_tier - 1) * 44

            # Extending end date only
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date
            
        except:
            print(f"Current Column {column} was unable to process...")

    # If type is "F" or "P", update prices in Pricing Grids for the respective tier
    elif pricing_type in ("F", "P"):
        try:  
            row_number = (136 + (current_tier - 1) * 44) + 44

            # Discounts, dates, and pricing applied
            values = [
                0, pricing_date, end_date,
                AI_OS, AI_VS, AI_OC,
                AI_PH1, AI_PH2, AI_PH3,
                AI_A1, AI_A2, AI_A3,
                AI_A4, AI_B1, AI_B2,
                AI_B3, AI_B4, AI_C,
                AI_F, AI_G
            ]

            for i, value in enumerate(values):
                pricing_grids_sheet.cell(row=row_number + i, column=pricing_grids_column).value = value

            # Increase the tier of price by 1
            pricing_grids_sheet.cell(row=66, column=pricing_grids_column).value = current_tier + 1

        except:
            print(f"Current Column {column} was unable to process...")

def A_Class_Change(Info):
    column, price_changes_sheet, pricing_grids_sheet, pricing_grids_column, end_date, pricing_date = Info[:6]

    # store pricing values
    AI_values = []
    row_start = 182
    row_end = 196

    for row in range(row_start, row_end + 1):
        AI_values.append(price_changes_sheet.cell(row=row, column=column).value)

    AI_OS, AI_VS, AI_OC, AI_PH1, AI_PH2, AI_PH3, AI_A1, AI_A2, AI_A3, AI_A4, AI_B1, AI_B2, AI_B3, AI_B4, AI_B5, AI_S = AI_values[:15]

    # Check the type in Price Changes row 85
    pricing_type = price_changes_sheet.cell(row=85, column=column).value

    # Check the Air Credit Action
    AirCreditAction = price_changes_sheet.cell(row=16, column=column).value

    # Current Tier
    current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value

    # If type is "N", extend the end date in Pricing Grids based on the tier
    if pricing_type == "N" and AirCreditAction == "FLAT":
        try:
            row_number = 132 + (current_tier - 1) * 42

            # Extending end date only
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date
            
        except:
            print(f"Current Column {column} was unable to process...")

    # If type is "F" or "P", update prices in Pricing Grids for the respective tier
    elif pricing_type in ("F", "P"):
        try:  
            row_number = (132 + (current_tier - 1) * 42) + 42

            # Discounts, dates, and pricing applied
            values = [
                0, pricing_date, end_date,
                AI_OS, AI_VS, AI_OC,
                AI_PH1, AI_PH2, AI_PH3,
                AI_A1, AI_A2, AI_A3,
                AI_A4, AI_B1, AI_B2,
                AI_B3, AI_B4, AI_B5, AI_S
            ]

            for i, value in enumerate(values):
                pricing_grids_sheet.cell(row=row_number + i, column=pricing_grids_column).value = value

            # Increase the tier of price by 1
            pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value = current_tier + 1

        except:
            print(f"Current Column {column} was unable to process...")

def Price_Changes(ArgInfo):
    PriceGridFile = ArgInfo[0]
    PriceChangeFile = ArgInfo[1]
    Ship = ArgInfo[2]
    pricing_date = ArgInfo[3]
    end_date = ArgInfo[4]
    start_cruise = ArgInfo[5]

    Ship_Dict = {
        "INS": ["Insignia","INS Details"],
        "NAU": ["Nautica","NAU Details"],
        "REG": ["Regatta","REG Details"],
        "SIR": ["Sirena","SIR Details"],
        "MNA": ["Marina","MNA Details"],
        "RVA": ["Riviera","RVA Details"],
        "VIS": ["Vista","VIS Details"]
    }

    if Ship in Ship_Dict.keys():
            Ship_Details = Ship_Dict[Ship]
    else:
        print(f"Ship Code was not recognized: {Ship}")
        return
    
    # Open the Price Change respective file
    price_changes_wb = openpyxl.load_workbook(PriceChangeFile, data_only=True)
    price_changes_sheet = price_changes_wb[Ship_Details[1]]

    # Find the inputed start_cruise in row 4 of Price Changes
    cruise_column = None
    for column in range(1, price_changes_sheet.max_column + 1):
        if price_changes_sheet.cell(row=4, column=column).value == start_cruise:
            cruise_column = column
            break
    
    # If the start_cruise entered is not found quit and notify user
    if cruise_column is None:
        print("start cruise not found in Price Changes.")
        return

    # Open the Pricing Grids file
    pricing_grids_wb = openpyxl.load_workbook(PriceGridFile, data_only=True)
    pricing_grids_sheet = pricing_grids_wb[Ship_Details[0]]

    # Find the column in Pricing Grids that corresponds to the start_cruise
    pricing_grids_column = None
    for column in range(1, pricing_grids_sheet.max_column + 1):
        if pricing_grids_sheet.cell(row=3, column=column).value == start_cruise:
            pricing_grids_column = column
            break
    
    # If the start_cruise entered is not found quit and notify user
    if pricing_grids_column is None:
        print("start_cruise not found in Pricing Grids.")
        return

    # Update prices and dates for the remaining cruises
    for column in range(cruise_column, price_changes_sheet.max_column + 1):
        Info = [column, price_changes_sheet, pricing_grids_sheet, pricing_grids_column, end_date, pricing_date]
        
        # Get NEW AC IN and NEW AC OUT values
        new_ac_in = price_changes_sheet.cell(row=14, column=column).value
        new_ac_out = price_changes_sheet.cell(row=15, column=column).value

        # Update NEW AC IN and NEW AC OUT in Pricing Grids
        pricing_grids_sheet.cell(row=18, column=pricing_grids_column).value = new_ac_in
        pricing_grids_sheet.cell(row=19, column=pricing_grids_column).value = new_ac_out

        # Changing price by category
        Class_Execution(Ship, Info)

        # Move to the next column in Price Changes and Pricing Grids
        cruise_column += 1
        pricing_grids_column += 1

    # Once completed save changes
    price_changes_wb.save(PriceChangeFile)
    pricing_grids_wb.save(PriceGridFile)

    # notify user of completion
    print(f"Pricing update completed successfully for {Ship}")

def main():
    # Standard user input
    Arg_Info = ArgSelector()
    Arg_Info = ArgConfirmation(Arg_Info)
    Price_Changes(Arg_Info)

if __name__ == "__main__":
    main()