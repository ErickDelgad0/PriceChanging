# Created by Erick Delgado June 8th 2023

# Formatted for INS Sailing Price Changes

# Dedicated to Owen and Jessica the grid slaying GODS

#             .-------.
#             |(~\o/~)|
#             ||\/X\/||
#             || \ / ||
#             ||o X o||
#             || / \ ||
#             ||/\X/\||
# .----------._)~   ~(_.----------.
# |\/)~~(\/\   (~\ /~)   /\/)~~(\/|
# |(X () X) >o  >-X-<  o< (X () X)|
# |/\)__(/\/  _(_/|\_)_  \/\)__(/\|
# '----------' )     ( '----------'
#             ||\/~\/||
#             || \o/ ||
#             |(~\ /~)|
#             ||\/X\/||
#             || \ / ||
#             ||  X  ||
#             ||\(/\/||
#             ||=)O(=||
#             ||/\/)\||
#             ||  X  ||
#             || / \ ||
#             ||/\X/\||
#             |(_/o\_)|
#             '._____.'
# -----------------------------------------------

import openpyxl

def INS_update_pricing(cruise, end_date, pricing_date):
    # Open the respective file
    price_changes_wb = openpyxl.load_workbook('Excel-Files/Pricing-Change.xlsx')
    price_changes_sheet = price_changes_wb['INS Details']

    # Find the inputed cruise in row 4 of Price Changes
    cruise_column = None
    for column in range(1, price_changes_sheet.max_column + 1):
        if price_changes_sheet.cell(row=4, column=column).value == cruise:
            cruise_column = column
            break
    
    # If the cruise entered is not found quit and notify user
    if cruise_column is None:
        print("Cruise not found in Price Changes.")
        return

    # Open the Pricing Grids file
    pricing_grids_wb = openpyxl.load_workbook('Excel-Files/Grid-of-doom.xlsx')
    pricing_grids_sheet = pricing_grids_wb['Insignia']

    # Find the column in Pricing Grids that corresponds to the cruise
    pricing_grids_column = None
    for column in range(1, pricing_grids_sheet.max_column + 1):
        if pricing_grids_sheet.cell(row=3, column=column).value == cruise:
            pricing_grids_column = column
            break
    
    # If the cruise entered is not found quit and notify user
    if pricing_grids_column is None:
        print("Cruise not found in Pricing Grids.")
        return

    # Update prices and dates for the remaining cruises
    for column in range(cruise_column, price_changes_sheet.max_column + 1):
        # Get NEW AC IN and NEW AC OUT values
        new_ac_in = price_changes_sheet.cell(row=14, column=column).value
        new_ac_out = price_changes_sheet.cell(row=15, column=column).value

        # Get the new all inclusive prices changes
        AI_OS = price_changes_sheet.cell(row=182, column=column).value
        AI_VS = price_changes_sheet.cell(row=183, column=column).value
        AI_PH1 = price_changes_sheet.cell(row=184, column=column).value
        AI_PH2 = price_changes_sheet.cell(row=185, column=column).value
        AI_PH3 = price_changes_sheet.cell(row=186, column=column).value
        AI_A1 = price_changes_sheet.cell(row=187, column=column).value
        AI_A2 = price_changes_sheet.cell(row=188, column=column).value
        AI_A3 = price_changes_sheet.cell(row=189, column=column).value
        AI_B1 = price_changes_sheet.cell(row=190, column=column).value
        AI_B2 = price_changes_sheet.cell(row=191, column=column).value
        AI_C1 = price_changes_sheet.cell(row=192, column=column).value
        AI_C2 = price_changes_sheet.cell(row=193, column=column).value
        AI_D = price_changes_sheet.cell(row=194, column=column).value
        AI_S = price_changes_sheet.cell(row=195, column=column).value
        AI_F = price_changes_sheet.cell(row=196, column=column).value
        AI_G = price_changes_sheet.cell(row=197, column=column).value

        # Update NEW AC IN and NEW AC OUT in Pricing Grids (shifted one cell/row upwards)
        pricing_grids_sheet.cell(row=17, column=pricing_grids_column).value = new_ac_in
        pricing_grids_sheet.cell(row=18, column=pricing_grids_column).value = new_ac_out

        # Check the type in Price Changes row 85
        pricing_type = price_changes_sheet.cell(row=85, column=column).value

        # If type is "N", extend the end date in Pricing Grids based on the tier
        if pricing_type == "N":
            current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value
            row_number = 132 + (current_tier - 1) * 42

            # Extending end date only
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date

        # If type is "F" or "P", update prices in Pricing Grids for the respective tier
        elif pricing_type in ("F", "P"):
            current_tier = pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value
            row_number = (132 + (current_tier-1) * 42) + 42

            # Discounts and dates applied
            pricing_grids_sheet.cell(row=row_number-3, column=pricing_grids_column).value = 0
            pricing_grids_sheet.cell(row=row_number-2, column=pricing_grids_column).value = pricing_date
            pricing_grids_sheet.cell(row=row_number-1, column=pricing_grids_column).value = end_date

            # New Pricing applied
            pricing_grids_sheet.cell(row=row_number+1, column=pricing_grids_column).value = AI_OS
            pricing_grids_sheet.cell(row=row_number+2, column=pricing_grids_column).value = AI_VS
            pricing_grids_sheet.cell(row=row_number+3, column=pricing_grids_column).value = AI_PH1
            pricing_grids_sheet.cell(row=row_number+4, column=pricing_grids_column).value = AI_PH2
            pricing_grids_sheet.cell(row=row_number+5, column=pricing_grids_column).value = AI_PH3
            pricing_grids_sheet.cell(row=row_number+6, column=pricing_grids_column).value = AI_A1
            pricing_grids_sheet.cell(row=row_number+7, column=pricing_grids_column).value = AI_A2
            pricing_grids_sheet.cell(row=row_number+8, column=pricing_grids_column).value = AI_A3
            pricing_grids_sheet.cell(row=row_number+9, column=pricing_grids_column).value = AI_B1
            pricing_grids_sheet.cell(row=row_number+10, column=pricing_grids_column).value = AI_B2
            pricing_grids_sheet.cell(row=row_number+11, column=pricing_grids_column).value = AI_C1
            pricing_grids_sheet.cell(row=row_number+12, column=pricing_grids_column).value = AI_C2
            pricing_grids_sheet.cell(row=row_number+13, column=pricing_grids_column).value = AI_D
            pricing_grids_sheet.cell(row=row_number+14, column=pricing_grids_column).value = AI_S
            pricing_grids_sheet.cell(row=row_number+15, column=pricing_grids_column).value = AI_F
            pricing_grids_sheet.cell(row=row_number+16, column=pricing_grids_column).value = AI_G

            # Increase the tier of price by 1
            pricing_grids_sheet.cell(row=65, column=pricing_grids_column).value = current_tier + 1

        # Move to the next column in Price Changes and Pricing Grids
        cruise_column += 1
        pricing_grids_column += 1

    # Once completed save changes
    price_changes_wb.save('Excel-Files/Pricing-Change.xlsx')
    pricing_grids_wb.save('Excel-Files/Grid-of-doom.xlsx')

    # notify user of completion
    print("Pricing update completed successfully for Insignia")


def main():
    # Standard user input
    cruise = input("Select the first cruise that the Pricing Date will take effect on: ")
    end_date = input("Enter the End Date: ")
    pricing_date = input("Enter the Pricing Date: ")

    # Perform the updates
    INS_update_pricing(cruise, end_date, pricing_date)

if __name__ == "__main__":
    main()